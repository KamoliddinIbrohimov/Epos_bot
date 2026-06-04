import io

import openpyxl
from aiogram import types
from aiogram.dispatcher.filters.builtin import CommandHelp

from handlers.users.bulk_block_dates import MAX_ROWS
from loader import dp


def _make_sample_xlsx() -> io.BytesIO:
    """Build an in-memory sample .xlsx illustrating the bulk-block-date
    format: column A = fiscal number (name), column B = block date."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Block dates"

    samples = [
        ("VG544170009490", "2026-08-01"),
        ("UZ210317264700", "01.09.2026"),
        ("LG420211628059", "2026/10/01"),
    ]
    for row_idx, (fn, dt) in enumerate(samples, start=1):
        ws.cell(row=row_idx, column=1, value=fn)
        ws.cell(row=row_idx, column=2, value=dt)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@dp.message_handler(CommandHelp(), chat_type=types.ChatType.PRIVATE)
async def bot_help(message: types.Message):
    text = (
        "📌 <b>Команды</b>\n"
        "/start — запуск бота\n"
        "/help — эта справка\n\n"
        "🔒 <b>Обновление даты блокировки (одиночное)</b>\n"
        "Пришли фискальный номер (например <code>LG420211628059</code>) "
        "обычным сообщением — бот покажет инфо о клиенте и календарь, "
        "выбери дату, бот применит её.\n\n"
        "📤 <b>Массовое обновление даты блокировки</b>\n"
        "Пришли <b>.xlsx</b>-файл прямо в чат (без всяких команд).\n"
        "Формат:\n"
        "• Колонка <b>A</b> — фискальный номер (поле <code>name</code> в Cazad)\n"
        "• Колонка <b>B</b> — дата блокировки\n\n"
        "Принимаемые форматы даты:\n"
        "<code>2026-08-01</code>, <code>2026/08/01</code>, "
        "<code>01.08.2026</code>, <code>01-08-2026</code>, "
        "<code>01/08/2026</code>, а также нативные даты Excel.\n\n"
        f"Лимит — <b>{MAX_ROWS}</b> строк за файл. Если больше — разбей на части.\n\n"
        "После загрузки бот покажет предпросмотр и попросит подтверждения. "
        "По «❌ Отмена» ничего не меняется.\n\n"
        "Ниже — пример файла, можно скачать как шаблон."
    )
    await message.answer(text)

    buf = _make_sample_xlsx()
    await message.answer_document(
        types.InputFile(buf, filename="block_dates_sample.xlsx"),
        caption="Пример формата для массового обновления.",
    )
